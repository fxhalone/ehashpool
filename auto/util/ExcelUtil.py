import openpyxl
import pandas as pd



def getRowsCount(filePath,sheetName):
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]
    maxRows = sh.max_row
    wb.close()
    return maxRows


def getAllColsData(filePath,sheetName,colNumber):

    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]
    maxRows = sh.max_row
    data = []
    for i in range(2,maxRows,1):
        data.append(sh.cell(i,colNumber).value)
    return data


def tianchong(filePath,sheetName):
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]
    a = 1
    for i in range(2,167076,1):
        sh.cell(i,1,a)
        a = a+1
    wb.save(filePath)


def compare(filePath1,sheetName1,filePath2,sheetName2):
    df = pd.read_excel(filePath1,sheetName1)
    c = df['市'].value_counts().to_dict()
    print(c)
    wb = openpyxl.load_workbook(filePath2)
    sh = wb[sheetName2]
    num = sh.max_row
    print(num)
    a = {}
    for i in range(2,num+1,1):
        a[sh.cell(i,2).value] = sh.cell(i,1).value
    print(a)

    flag = True
    for key,value in c.items():
        if flag == False:
            break
        for key1, value1 in a.items():
            if key == key1:
                if value!=value1:
                    print(key,value,key1,value1)
                    flag = False
                    break
    print(flag)
    print(set(c.keys()) ^ set(a.keys()))









if __name__ == '__main__':


    # wb = openpyxl.load_workbook("/Users/fengxiaohui/Desktop/数据/电池故障导入模版.xlsx")
    # sh = wb["Sheet1"]
    # for i in range(2,488547,1):
    #     sh.cell(i,1,i-1)
    #
    #
    #
    # wb.save("/Users/fengxiaohui/Desktop/数据/电池故障导入模版.xlsx")
    # pass
    # print(getAllColsData("/Users/fengxiaohui/Desktop/10万组电池发货导入.xlsx","Sheet1",1))
    # tianchong("/Users/fengxiaohui/Desktop/故障倒入.xlsx","Sheet1")
    # compare("/Users/fengxiaohui/Desktop/zoubao1.xlsx","Sheet0","/Users/fengxiaohui/Desktop/无标题.xlsx","Sheet1")
    #
    # wb = openpyxl.load_workbook("/Users/fengxiaohui/Desktop/数据/电池故障导入模版.xlsx")
    #
    # df = pd.read_excel("/Users/fengxiaohui/Desktop/数据/123.xlsx", "Sheet0")
    # c = df['市'].value_counts().to_dict()
    # print(c)
    # print(325057+319+31+23+18+13+3)
    #
    # print(163488+325057)
    getRowsCount()



    #
    # wb = openpyxl.load_workbook("/Users/fengxiaohui/Desktop/报表/2.xlsx")
    # sh = wb["Sheet0"]
    # num = sh.max_row
    # print(num)
    # max = 0
    # for i in range(2,num+1,1):
    #     print(sh.cell(i,3).value)
    #     max += int(sh.cell(i,3).value)
    #
    #
    # print(max+3276)
    # # print(158616+8490)
    #
    # df = pd.read_excel("/Users/fengxiaohui/Desktop/数据/123.xlsx", "Sheet0")
    # e = df['市'].value_counts().to_dict()
    # print(e)
    # a = "沈阳市 | 41440 | | 银川市 | 11453 | | 阜阳市 | 9176 | | 拉萨市 | 7762 | | 宜春市 | 7305 | | 商丘市 | 6621 | | 梅州市 | 5060 | | 呼和浩特市 | 4632 | | 乌兰察布市 | 4379 | | 开封市 | 4333 | | 郑州市 | 4203 | | 乌海市 | 4138 | | 西双版纳傣族自治州 | 4079 | | 茂名市 | 4027 | | 鄂尔多斯市 | 3785 | | 日喀则市 | 3106 | | 大连市 | 3034 | | 德州市 | 2856 | | 阜新市 | 2839 | | 淄博市 | 2556 | | 伊犁哈萨克自治州 | 2510 | | 新疆自治区直辖县 | 2332 | | 漳州市 | 1575 | | 营口市 | 1523 | | 汉中市 | 1380 | | 咸阳市 | 1272 | | 吉林市 | 1142 | | 十堰市 | 1049 | | 南宁市 | 981 | | 信阳市 | 886 | | 通辽市 | 810 | | 南昌市 | 695 | | 唐山市 | 693 | | 合肥市 | 571 | | 鸡西市 | 570 | | 汕头市 | 516 | | 济南市 | 516 | | 六安市 | 494 | | 淮北市 | 477 | | 长春市 | 419 | | 南京市 | 401 | | 玉林市 | 376 | | 抚州市 | 353 | | NULL | 327 | | 新乡市 | 308 | | 湖州市 | 302 | | 上饶市 | 292 | | 临沂市 | 257 | | 菏泽市 | 230 | | 塔城地区 | 219 | | 镇江市 | 201 | | 洛阳市 | 201 | | 商洛市 | 192 | | 巴彦淖尔市 | 183 | | 温州市 | 173 | | 徐州市 | 150 | | 金华市 | 141 | | 大庆市 | 137 | | 保定市 | 135 | | 本溪市 | 133 | | 揭阳市 | 123 | | 吉安市 | 122 | | 宁德市 | 122 | | 省直辖县级行政区划 | 117 | | 杭州市 | 94 | | 九江市 | 93 | | 赤峰市 | 90 | | 酒泉市 | 86 | | 包头市 | 73 | | 台州市 | 65 | | 吐鲁番市 | 62 | | 亳州市 | 60 | | 桂林市 | 49 | | 厦门市 | 48 | | 永州市 | 43 | | 济宁市 | 40 | | 赣州市 | 37 | | 蚌埠市 | 32 | | 昆明市 | 30 | | 榆林市 | 28 | | NULL | 28 | | 周口市 | 23 | | 清远市 | 19 | | 丽水市 | 18 | | 石家庄市 | 18 | | 绍兴市 | 17 | | 贵港市 | 14 | | 石嘴山市 | 13 | | 新余市 | 13 | | 宿州市 | 12 | | 楚雄彝族自治州 | 12 | | 焦作市 | 11 | | 海口市 | 11 | | 三门峡市 | 10 | | 乐山市 | 8 | | 宣城市 | 8 | | 嘉兴市 | 7 | | 烟台市 | 7 | | 武汉市 | 6 | | 成都市 | 5 | | 松原市 | 5 | | 张掖市 | 5 | | 三明市 | 5 | | 昌吉回族自治州 | 4 | | 沧州市 | 3 | | 哈尔滨市 | 3 | | NULL | 3 | | 西安市 | 2 | | 钦州市 | 2 | | 黄冈市 | 2 | | 阿坝藏族羌族自治州 | 1 | | 龙岩市 | 1 | | 长沙市 | 1 | | 阿拉善盟 | 1 | | 海东市 | 1 | | 泉州市 | 1 | | 无锡市 | 1 | | 梧州市 | 1 | | 延安市 | 1 | | 邢台市 | 1 | | 马鞍山市 | 1 | | 安庆市 | 1 | | 南阳市 | 1 | | 滁州市 | 1 | | 晋城市 | 1 | | 枣庄市 | 1 | | 河源市 | 1 | | 巴音郭楞蒙古自治州 | 1 |"
    # b = a.replace(' ','')
    # c = b.replace('||','|')
    # d = c[0:-1]
    # list1 = d.split('|')
    # # print(len(list1))
    # map1 = {}
    # for i in range(0,int(len(list1)/2)+1,2):
    #     map1[list1[i]] = int(list1[i+1])
    #
    # print(map1)
    #
    #
    # flag = True
    # for key, value in e.items():
    #     # if flag == False:
    #     #     break
    #     for key1, value1 in map1.items():
    #         if key == key1:
    #             if value != value1:
    #                 print(key, value, key1, value1)
    #                 # flag = False
    #                 # break
    # # print(flag)
